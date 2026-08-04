"""
Schedule Parser Module (schedule_parser.py)
AI-powered PDF/text analyzer for Korean martial arts training schedules.
Extracts annual and monthly event data and maps them to calendar dates.
"""

import os
import json
import re
import calendar
from datetime import datetime, date, timedelta
from typing import Optional

# PDF extraction
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False
    print("⚠️ pdfplumber가 없습니다. pip install pdfplumber 실행 후 사용하세요.")

# Excel manipulation
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.cell.text import InlineFont
    from openpyxl.worksheet.pagebreak import Break
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl이 없습니다. pip install openpyxl 실행 후 사용하세요.")


def clean_training_text(text: str) -> str:
    """
    Convert delimiters (/ or ,) into newlines for better readability.
    Also strips extra whitespace.
    """
    if not text:
        return ""
    # Replace slashes and commas with newlines, then strip/clean
    # re.sub finds both / and ,
    cleaned = re.sub(r'[/,]', '\n', text)
    # Filter empty lines and strip each line
    lines = [line.strip() for line in cleaned.split('\n') if line.strip()]
    
    # 3번째 라인(주 기술) 강조 (사용자 요청 반영: 눈에 잘 뛰게)
    if len(lines) >= 3:
        # 이미 강조 표시가 없는 경우에만 추가
        if not lines[2].startswith("▶"):
            lines[2] = f"▶ {lines[2]}"
            
    return '\n'.join(lines)


def get_smart_alignment(text):
    """
    텍스트 내용에 따라 적절한 Alignment를 반환합니다.
    - 줄바꿈이 있는 경우: wrapText 우선 (shrink_to_fit은 wrap과 함께 쓸 수 없음)
    - 줄바꿈 없이 긴 텍스트: shrink_to_fit 적용 (글자 수에 따라 자동으로 작아짐)
    """
    has_newline = isinstance(text, str) and '\n' in text
    if has_newline:
        return Alignment(wrapText=True, vertical='center', horizontal='center')
    else:
        return Alignment(shrink_to_fit=True, vertical='center', horizontal='center')



def get_dynamic_font_size(text: str):
    """
    텍스트 길이에 따라 적절한 폰트 크기를 반환합니다. (7~9pt)
    """
    if not text: return 9
    text_str = text if isinstance(text, str) else str(text)
    # 글자 수가 많아지면 폰트 크기를 단계적으로 축소
    if len(text_str) > 40: return 7
    if len(text_str) > 20: return 8
    return 9


def format_cell_text(text: str, period: str = "오후", is_annual: bool = False):
    """
    엑셀 호환성을 위해 복잡한 스타일 대신 일반 텍스트를 반환합니다.
    줄바꿈만 정돈하여 깔끔하게 표시되도록 합니다.
    """
    if not text:
        return ""
    return clean_training_text(text)


def extract_text_from_pdf(pdf_path: str) -> str:
    """
    Extract text content from a PDF file.
    Returns concatenated text from all pages.
    """
    if not PDFPLUMBER_AVAILABLE:
        raise ImportError("pdfplumber 라이브러리가 필요합니다. pip install pdfplumber")

    text_parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Extract regular text
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
            
            # Also try to extract tables
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if row:
                        cleaned_row = [str(cell).strip() if cell else "" for cell in row]
                        text_parts.append(" | ".join(cleaned_row))
    
    return "\n".join(text_parts)


def extract_text_from_file(file_path: str, openai_client) -> str:
    """
    파일 확장자에 따라 이미지(Vision) 또는 PDF 텍스트를 추출합니다.
    """
    ext = file_path.lower().split('.')[-1]
    if ext in ['pdf']:
        print(f"📄 PDF 텍스트 추출 중: {file_path}")
        return extract_text_from_pdf(file_path)
    elif ext in ['png', 'jpg', 'jpeg']:
        print(f"🖼️ 이미지(Vision) 텍스트 추출 중: {file_path}")
        import base64
        try:
            with open(file_path, "rb") as image_file:
                base64_image = base64.b64encode(image_file.read()).decode('utf-8')
            
            response = openai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": "이 이미지에 적힌 모든 행사, 일정, 수련 계획표 텍스트를 하나도 빠짐없이 정확하게 텍스트로 추출해주세요."},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/{ext};base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                max_tokens=4000
            )
            extracted_text = response.choices[0].message.content
            print(f"✅ 이미지 텍스트 추출 완료 ({len(extracted_text)}자)")
            return extracted_text
        except Exception as e:
            print(f"❌ 이미지 텍스트 추출 오류: {e}")
            return ""
    else:
        # 일반 텍스트 파일 읽기
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            print(f"❌ 파일 읽기 오류: {e}")
            return ""


def parse_schedule_with_gpt(raw_text: str, year: int, mode: str, ai_handler=None, openai_client=None) -> list[dict]:
    """
    AI를 사용하여 PDF 텍스트에서 구조화된 일정 데이터를 추출합니다.
    AIHandler가 제공되면 이를 우선 사용하고, 없으면 기존 openai_client를 사용합니다.
    """
    print(f"🔍 [AI] {mode} 일정 추출 시작 (텍스트 길이: {len(raw_text)})")
    
    if mode == "annual":
        system_prompt = f"""당신은 무도 협회(태권도, 합기도 등)의 연간 일정표를 분석하는 최고의 전문가입니다.
추출된 텍스트에서 모든 행사, 대회, 교육 일정을 정확하게 JSON 배열로 추출하세요.

[필수 규칙]
1. 연도 설정: 문서 내에 명시된 연도(예: 2024년, 2025년)를 최우선으로 따르세요. 만약 문서에 연도 정보가 전혀 없다면 {year}년을 기준으로 삼으세요.
2. 날짜 범위 처리: '3월 12~15일'과 같은 범위는 마지막 종료 날짜(15일)를 절대 빼먹지 말고, 반드시 2026-03-12, 2026-03-13, 2026-03-14, 2026-03-15 처럼 시작일부터 종료일까지 하루도 빠짐없이 모든 개별 날짜를 분리해서 각각의 항목으로 생성해야 합니다.
3. 날짜 형식: 반드시 'YYYY-MM-DD' 형식을 지키세요.
4. 항목 누락 금지: 텍스트에 있는 모든 공식 일정을 빠짐없이 추출하세요.
5. 수련단계(stage): 정보가 없으면 기본값인 3(기술)을 사용하세요. (1:준비, 2:체력, 3:기술, 4:마무리)

JSON 형식: [{{"date": "YYYY-MM-DD", "title": "내용", "location": "장소", "stage": 3}}]
"""
    else:  # monthly
        system_prompt = f"""당신은 태권도/무도 수련 계획표를 분석하는 전문가입니다.
날짜별 수련 내용을 상세히 추출하여 JSON 배열로 만드세요.

[필수 규칙]
1. 연월 감지: 문서 상단의 'X년 X월 수련계획' 정보를 가장 먼저 찾아 해당 연도와 월을 기준으로 삼으세요. (과거 연도라도 문서에 적힌 대로 추출하세요)
2. 줄바꿈 유지: 수련 내용이 여러 개면 반드시 '\\n' 문자로 구분하세요.
3. 정확성: 다른 설명 없이 오직 JSON 배열만 출력하세요.

JSON 형식: [{{"date": "YYYY-MM-DD", "title": "내용1\\n내용2", "location": "", "stage": 3}}]
"""

    user_message = f"다음 텍스트에서 모든 일정을 추출해 JSON으로 응답해줘:\n\n{raw_text[:6000]}"

    try:
        if ai_handler and hasattr(ai_handler, 'ask'):
            # AIHandler 사용 (Gemini 등 지원)
            raw_response = ai_handler.ask(
                user_prompt=user_message,
                system_prompt=system_prompt,
                max_tokens=4000
            )
        elif openai_client:
            # 기존 OpenAI 클라이언트 사용
            response = openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_message}
                ],
                temperature=0.0,
                response_format={ "type": "json_object" } if mode == "annual" else None
            )
            raw_response = response.choices[0].message.content
        else:
            print("❌ [AI] 호출 가능한 AI 클라이언트가 없습니다.")
            return []
        print(f"📡 [AI 응답 수신완료]")
        
        # JSON 파싱 (json_object 모드일 경우 루트 키가 있을 수 있음)
        try:
            data = json.loads(raw_response)
            if isinstance(data, dict):
                # 'events', 'schedule' 등 키가 있을 경우 리스트만 추출
                for key in data:
                    if isinstance(data[key], list):
                        entries = data[key]
                        break
                else:
                    entries = []
            else:
                entries = data
        except:
            # 텍스트에서 JSON 배열만 정규식으로 추출 시도
            import re
            match = re.search(r'\[.*\]', raw_response, re.DOTALL)
            if match:
                entries = json.loads(match.group())
            else:
                entries = []
        
        print(f"✅ AI가 {len(entries)}개의 일정을 추출했습니다.")
        return entries

    except Exception as e:
        print(f"❌ AI 파싱 오류: {e}")
        return []


def write_annual_schedule_to_excel(excel_path: str, schedule_entries: list[dict]) -> bool:
    """연간 일정 목록을 엑셀 파일(연간계획 시트)에 기록 (템플릿 생성 포함)"""
    import unicodedata
    excel_path = unicodedata.normalize('NFC', excel_path)
    
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl이 필요합니다. pip install openpyxl")

    try:
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # Ensure '연간계획' sheet exists
        if "연간계획" not in wb.sheetnames:
            ws = wb.create_sheet("연간계획", 0)
            _init_annual_sheet(ws)
        else:
            ws = wb["연간계획"]

        # 1.5. 월별 시트가 없으면 1~12월까지 모두 생성
        current_year = datetime.now().year
        for m in range(1, 13):
            m_name = f"{m}월"
            if m_name not in wb.sheetnames:
                m_ws = wb.create_sheet(m_name)
                _init_monthly_sheet(m_ws, current_year, m)

        # Find the actual last row with data in column A
        last_row = 2
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            if row[0].value is not None:
                last_row = row[0].row
        
        start_row = last_row + 1

        # Write entries
        fill_colors = {
            "오전": PatternFill("solid", fgColor="FFF8F0"),
            "오후": PatternFill("solid", fgColor="F0F8FF"),
            "저녁": PatternFill("solid", fgColor="F8F8FF"),
        }

        def get_event_color(title_text):
            if any(k in title_text for k in ["심사", "승급", "테스트"]):
                return "FF8B00FF" # 보라색
            elif any(k in title_text for k in ["대회", "시합", "체전", "연습"]):
                return "FF008000" # 녹색
            elif any(k in title_text for k in ["특강", "세미나", "교육", "캠프", "체험", "나들이", "합숙", "[", "]"]):
                return "FF0000FF" # 파란색
            return "FF000000"

        for i, entry in enumerate(schedule_entries):
            row_num = start_row + i
            title_only = entry.get("title", entry.get("event", ""))
            
            # 이벤트 타입에 따른 글씨 색상 결정
            font_color = get_event_color(title_only)
            event_font = Font(bold=True, color=font_color)

            ws.cell(row=row_num, column=1, value=entry.get("date", "")).font = event_font
            ws.cell(row=row_num, column=2, value=title_only).font = event_font
            ws.cell(row=row_num, column=3, value=entry.get("location", "")).font = event_font
            ws.cell(row=row_num, column=4, value=entry.get("stage", "3")).font = event_font
            ws.cell(row=row_num, column=5, value="기본 일정").font = event_font

            stage = str(entry.get("stage", "3"))
            # 배경색은 단계별로 살짝 다르게 (선택 사항)
            fill = fill_colors.get("오후", fill_colors["오후"]) 
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).fill = fill
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # --- 추가: 각 월별 시트(달력 칸)에도 '적용 가능할 때' 기록 ---
        monthly_map = {}
        for entry in schedule_entries:
            date_str = entry.get("date", "")
            if not date_str: continue
            try:
                m = datetime.strptime(date_str, "%Y-%m-%d").month
                month_key = f"{m}월"
                if month_key not in monthly_map: monthly_map[month_key] = []
                monthly_map[month_key].append(entry)
            except: continue

        for month_name, month_entries in monthly_map.items():
            if month_name in wb.sheetnames:
                m_ws = wb[month_name]
                for entry in month_entries:
                    date_str = entry.get("date", "")
                    raw_title = entry.get("title", entry.get("event", ""))
                    location = entry.get("location", "")
                    if not raw_title: continue

                    title = f"{raw_title}\n{location}" if location else raw_title
                    f_color = get_event_color(raw_title)
                    
                    for row in m_ws.iter_rows(min_row=31, max_row=m_ws.max_row):
                        cell_val = row[0].value
                        if not cell_val: continue
                        
                        if isinstance(cell_val, datetime):
                            cell_date_str = cell_val.strftime("%Y-%m-%d")
                        else:
                            cell_date_str = str(cell_val).strip()[:10]
                            
                        if cell_date_str == date_str or cell_date_str.endswith(date_str[5:]):
                            # 연간 행사는 기본적으로 '3단계 (기술)' 칸의 수식을 따라가서 본문에 기록
                            target_machine_cell = row[3] # 3단계 (D열)
                            
                            def get_fs(text):
                                l = len(str(text))
                                if l < 10: return 12
                                elif l < 20: return 10
                                else: return 9
                                
                            fs = get_fs(title)
                            
                            if target_machine_cell.value and str(target_machine_cell.value).startswith("="):
                                target_address = str(target_machine_cell.value).replace("=", "").strip()
                                try:
                                    target_cell = m_ws[target_address]
                                    # 데이터 보호: 빈칸일 때만 기록
                                    if not target_cell.value:
                                        target_cell.value = title
                                        target_cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                                        target_cell.font = Font(size=fs, bold=True, color=f_color)
                                except Exception as e: 
                                    print(f"달력 타겟 기록 실패: {e}")
                            else:
                                if not target_machine_cell.value:
                                    target_machine_cell.value = title
                                    target_machine_cell.font = Font(size=fs, bold=True, color=f_color)
                            break

        wb.save(excel_path)
        if len(schedule_entries) > 0:
            print(f"✅ '{excel_path}'에 {len(schedule_entries)}개 일정이 저장되었습니다.")
        else:
            print(f"✅ '{excel_path}'에 빈 수련계획표 템플릿 양식이 생성되었습니다.")
        return True
    except Exception as e:
        print(f"❌ 엑셀 기록 오류: {e}")
        return False



def get_holidays_for_month(y, m):
    """표준 달력 데이터 및 선거일 데이터를 포함한 해당 월의 공휴일 목록을 반환합니다."""
    kr_holidays = {}
    try:
        import holidays
        kr_holidays = holidays.SouthKorea(years=y)
    except Exception as err:
        print(f"⚠️ holidays 모듈 사용 불가 (기본 공휴일 맵으로 우회): {err}")
        # 기본 고정 공휴일 맵
        static_kr_holidays = {
            (1, 1): "신정",
            (3, 1): "삼일절",
            (5, 5): "어린이날",
            (6, 6): "현충일",
            (8, 15): "광복절",
            (10, 3): "개천절",
            (10, 9): "한글날",
            (12, 25): "성탄절"
        }
        for (sm, sd), hname in static_kr_holidays.items():
            if sm == m:
                kr_holidays[f"{y}-{sm:02d}-{sd:02d}"] = hname
    
    # [정밀 데이터] 법령 및 선거일 수동 반영 (표준 달력 참조)
    special_days = {
        2026: {
            (3,2): "대체공휴일(3·1절)",
            (5,25): "대체공휴일(부처님오신날)",
            (6,3): "제9회 지방선거",
            (8,17): "대체공휴일(광복절)",
            (10,5): "대체공휴일(개천절)"
        },
        2027: {
            (2,9): "대체공휴일(설날)",
            (3,3): "제21대 대통령선거",
            (8,16): "대체공휴일(광복절)",
            (10,4): "대체공휴일(개천절)",
            (10,11): "대체공휴일(한글날)",
            (12,27): "대체공휴일(성탄절)"
        }
    }
    
    month_holidays = {}
    # 해당 월의 공휴일만 필터링
    for date, name in kr_holidays.items():
        if date.year == y and date.month == m:
            h_name = name
            # 영문 명칭을 표준 한글 명칭으로 매핑
            if "New Year's Day" in name: h_name = "신정"
            elif "Independence Movement Day" in name: h_name = "삼일절"
            elif "Children's Day" in name: h_name = "어린이날"
            elif "Memorial Day" in name: h_name = "현충일"
            elif "Liberation Day" in name: h_name = "광복절"
            elif "National Foundation Day" in name: h_name = "개천절"
            elif "Hangeul Day" in name: h_name = "한글날"
            elif "Christmas" in name or "기독" in name: h_name = "성탄절"
            elif "Lunar New Year" in name: h_name = "설날"
            elif "Chuseok" in name: h_name = "추석"
            elif "Buddha's Birthday" in name: h_name = "부처님오신날"
            elif "Alternative" in name or "Substitute" in name: h_name = "대체공휴일"

            
            # [변경] 키를 숫자(day)가 아닌 표준 날짜 문자열(YYYY-MM-DD)로 저장
            date_key = f"{y}-{m:02d}-{date.day:02d}"
            month_holidays[date_key] = h_name
            
    # [중요] 수동 등록된 정밀 데이터로 덮어쓰기 (표준성 보장)
    if y in special_days:
        for (sm, sd), sname in special_days[y].items():
            if sm == m:
                date_key = f"{y}-{sm:02d}-{sd:02d}"
                month_holidays[date_key] = sname
                
    return month_holidays



def fix_layout_and_holidays(ws, year, month):
    """시트의 레이아웃(요일 헤더), 파스텔 테마 디자인, 공휴일 및 아이콘을 적용합니다."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    try:
        from openpyxl.drawing.image import Image as XLImage
    except ImportError as e:
        print(f"⚠️ 엑셀 이미지 삽입 모듈(Pillow) 없음 - 이미지 없이 진행합니다: {e}")
        XLImage = None
    import os

    # 테두리 스타일 정의 (구글 스프레드시트 호환을 위해 투명도 FF 추가)
    thin_side = Side(style='thin', color='FF000000')
    medium_side = Side(style='medium', color='FF000000')
    
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=thin_side, right=thin_side, top=medium_side, bottom=medium_side)

    # 0. 프린트 및 용지 설정 (A4 가로 꽉 차게)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    
    # [너비 설정] A4 가로 기준 7개 열이 꽉 차도록 설정 (너비 20~22)
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 22

    # 1. 파스텔 테마 색상 정의 (구글 호환 AARRGGBB)
    fill_sun = PatternFill("solid", fgColor="FFFFF0F5") # 연한 핑크
    fill_sat = PatternFill("solid", fgColor="FFE6E6FA") # 연한 라벤더
    fill_weekday = PatternFill("solid", fgColor="FFF0F8FF") # 연한 하늘
    fill_empty = PatternFill("solid", fgColor="FFF5F5F5") # 연회색
    fill_header = PatternFill("solid", fgColor="FFE0F2F1") # 연민트
    
    # 텍스트 이모지 매핑 (이미지 파일 없을 때 대비)
    holiday_emojis = {
        "신정": "☀️", "설날": "🙇", "삼일절": "🇰🇷", "어린이날": "🎁", 
        "부처님오신날": "🏮", "현충일": "🇰🇷", "광복절": "🇰🇷", "추석": "🎑", 
        "개천절": "🇰🇷", "한글날": "📝", "성탄절": "🎄", "지방선거": "🗳️"
    }

    # 2. 타이틀 및 요일 헤더 스타일 (Row 1 & 2)
    # [타이틀] 🦁 라이온짐 타이틀 강조 (Row 1)
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    if not title_cell.value:
        title_cell.value = f"🦁 {year}년 {month}월 라이온짐 수련계획표"
    title_cell.font = Font(size=18, bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="FFFFFF") # 흰색 배경

    # [요일 헤더] 스타일 (row 2)
    days = ["일", "월", "화", "수", "목", "금", "토"]
    for i, day in enumerate(days, 1):
        cell = ws.cell(row=2, column=i)
        cell.value = day
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill_header
        cell.border = header_border
        if i == 1: cell.font = Font(bold=True, color="FF4081") # 일요일 핫핑크
        elif i == 7: cell.font = Font(bold=True, color="3F51B5") # 토요일 진남색
        else: cell.font = Font(bold=True, color="00796B") # 평일 진초록


    # 3. 공휴일 정보 로드
    month_holidays = get_holidays_for_month(year, month)
    icon_dir = "resources/icons"

    # 4. 달력 그리드 디자인 적용 (row 3~29)
    for row_idx in range(3, 30, 2): # 날짜행과 내용행을 세트로 처리
        for col_idx in range(1, 8):
            date_cell = ws.cell(row=row_idx, column=col_idx)
            content_cell = ws.cell(row=row_idx + 1, column=col_idx)
            
            # 테두리 적용
            date_cell.border = thin_border
            content_cell.border = thin_border
            
            val = date_cell.value

            # 날짜가 없는 칸은 테두리와 배경색 제거 (디자인 깔끔하게)
            if not val or not isinstance(val, int):
                date_cell.fill = PatternFill(fill_type=None)
                content_cell.fill = PatternFill(fill_type=None)
                date_cell.border = Border()
                content_cell.border = Border()
                continue

            # 기본 배경색 (일/토/평일)
            if col_idx == 1: 
                date_cell.fill = fill_sun
                content_cell.fill = fill_sun
            elif col_idx == 7:
                date_cell.fill = fill_sat
                content_cell.fill = fill_sat
            else:
                date_cell.fill = fill_weekday
                content_cell.fill = fill_weekday

            # 공휴일 체크 (키 형식을 YYYY-MM-DD 문자열로 변환하여 비교)
            date_key = f"{year}-{month:02d}-{val:02d}"
            is_holiday = date_key in month_holidays
            if is_holiday or col_idx == 1:
                date_cell.font = Font(bold=True, color="FF0000")
                if is_holiday:
                    h_name = month_holidays[date_key]
                    
                    # 텍스트 이모지 접두사 찾기
                    prefix = ""
                    for key, emoji in holiday_emojis.items():
                        if key in h_name:
                            prefix = f"{emoji} "
                            break

                    if not content_cell.value or len(str(content_cell.value)) < 10:
                        content_cell.value = f"{prefix}{h_name}"
                    
                    content_cell.font = Font(bold=True, color="FF0000")
                    content_cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")


                    
                    # 공휴일 아이콘 삽입
                    icon_path = None
                    if "설날" in h_name: icon_path = os.path.join(icon_dir, "seollal.png")
                    elif "추석" in h_name: icon_path = os.path.join(icon_dir, "chuseok.png")
                    elif "어린이날" in h_name: icon_path = os.path.join(icon_dir, "childrens_day.png")
                    elif "성탄절" in h_name: icon_path = os.path.join(icon_dir, "xmas.png")
                    elif "선거" in h_name: icon_path = os.path.join(icon_dir, "election.png")
                    
                    if icon_path and os.path.exists(icon_path):
                        try:
                            if XLImage is not None:
                                img = XLImage(icon_path)
                                img.width, img.height = 40, 40 # 작게 조절
                                ws.add_image(img, content_cell.coordinate)
                        except: pass

            # 외발자전거 등 특강 아이콘 (사용자 요청)
            if content_cell.value and "외발자전거" in str(content_cell.value):
                icon_path = os.path.join(icon_dir, "unicycle.png")
                if os.path.exists(icon_path):
                    try:
                        if XLImage is not None:
                            img = XLImage(icon_path)
                            img.width, img.height = 35, 35
                            ws.add_image(img, content_cell.coordinate)
                    except: pass
            
            # 셀 높이 조절 (내용이 잘 보위도록)
            ws.row_dimensions[row_idx].height = 25
            ws.row_dimensions[row_idx + 1].height = 65

    # 5. 하단 기계 판독 영역 (31행~) 테두리 제거 (디자인 깔끔하게)
    from openpyxl.styles import Border, Side
    no_border = Border()
    for r in range(31, ws.max_row + 1):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = no_border


def update_excel_with_ai_results(excel_path: str, month: int, schedule_entries: list) -> bool:
    """
    AI가 생성한 일정 목록을 엑셀 파일의 해당 월 시트에 등록합니다.
    """
    if not OPENPYXL_AVAILABLE:
        print("⚠️ openpyxl이 설치되어 있지 않습니다.")
        return False

    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet_name = f"{month}월"
        
        # 파일 경로에서 연도 추출 (파일명에 20xx 포함 가정, 없으면 2026)
        year = 2026
        try:
            match = re.search(r"20\d{2}", os.path.basename(excel_path))
            if match: year = int(match.group())
        except: pass

        if sheet_name not in wb.sheetnames:
            print(f"⚠️ {sheet_name} 시트가 없습니다. 새로 생성합니다.")
            ws = wb.create_sheet(sheet_name)
            _init_monthly_sheet(ws, year, month)
        else:
            ws = wb[sheet_name]

        # --- [Step 1: 달력 초기화] ---
        # AI 결과 및 수동 기록을 지우되 공휴일은 재적용으로 복구
        for row in ws.iter_rows(min_row=31, max_row=ws.max_row):
            target_machine_cell = row[3] 
            if target_machine_cell.value and str(target_machine_cell.value).startswith("="):
                target_address = str(target_machine_cell.value).replace("=", "").strip()
                try:
                    ws[target_address].value = ""
                except: pass
            else:
                target_machine_cell.value = ""

        # 기존/신규 시트 공통 레이아웃 보정 및 공휴일 색상/텍스트 적용 (초기화 후 복구)
        fix_layout_and_holidays(ws, year, month)

        # --- [Step 2: 연간계획표 데이터 추출 (1순위)] ---
        all_entries = []
        annual_sheet_name = "연간계획"
        if annual_sheet_name in wb.sheetnames:
            annual_ws = wb[annual_sheet_name]
            for r in range(3, annual_ws.max_row + 1):
                a_date = annual_ws.cell(row=r, column=1).value
                a_title = annual_ws.cell(row=r, column=2).value
                a_loc = annual_ws.cell(row=r, column=3).value
                if a_date and a_title:
                    try:
                        if isinstance(a_date, datetime):
                            a_d = a_date
                        else:
                            a_d = datetime.strptime(str(a_date).strip()[:10].replace(".", "-"), "%Y-%m-%d")
                        if a_d.year == year and a_d.month == month:
                            # 장소가 있으면 합침
                            full_title = f"{a_title}\n{a_loc}" if a_loc else str(a_title)
                            all_entries.append({"date": a_d.strftime("%Y-%m-%d"), "title": full_title, "is_annual": True})
                    except: pass
        
        # --- [Step 3: AI 월간 진도 통합 (3순위)] ---
        for entry in schedule_entries:
            entry["is_annual"] = False
            all_entries.append(entry)

        # AI 및 연간 결과물 기록
        for entry in all_entries:
            date_str = entry.get("date", "")
            title = entry.get("title", entry.get("event", ""))
            is_annual = entry.get("is_annual", False)

            try:
                target_d = datetime.strptime(date_str, "%Y-%m-%d")
                target_md = target_d.strftime("%m-%d")
            except:
                target_md = None

            for row in ws.iter_rows(min_row=31, max_row=ws.max_row):
                cell_val = row[0].value
                
                if cell_val and str(cell_val).startswith("="):
                    try:
                        addr = str(cell_val).replace("=", "").strip()
                        cell_val = ws[addr].value
                    except: pass

                if not cell_val:
                    cell_md = ""
                else:
                    try:
                        if isinstance(cell_val, datetime):
                            cell_d = cell_val
                        else:
                            cell_v_str = str(cell_val).strip()[:10].replace(".", "-")
                            cell_d = datetime.strptime(cell_v_str, "%Y-%m-%d")
                        cell_md = cell_d.strftime("%m-%d")
                    except:
                        cell_md = str(cell_val)

                if cell_md == target_md and target_md is not None:
                    target_machine_cell = row[3] 
                    formatted_val = format_cell_text(title, "3단계")
                    text_for_style = formatted_val if isinstance(formatted_val, str) else title
                    smart_align = get_smart_alignment(text_for_style)
                    font_size = get_dynamic_font_size(text_for_style)
                    
                    def _get_event_color(t_text):
                        if any(k in t_text for k in ["심사", "승급", "테스트"]): return "8B00FF"
                        elif any(k in t_text for k in ["대회", "시합", "체전", "연습"]): return "008000"
                        elif any(k in t_text for k in ["특강", "세미나", "교육"]): return "0000FF"
                        elif any(k in t_text for k in ["캠프", "체험", "나들이", "합숙"]): return "FF8C00"
                        return "000000"
                        
                    f_color = _get_event_color(title) if is_annual else "000000"
                    
                    if target_machine_cell.value and str(target_machine_cell.value).startswith("="):
                        target_address = str(target_machine_cell.value).replace("=", "").strip()
                        try:
                            target_cell = ws[target_address]
                            
                            if is_annual:
                                # 1순위: 연간계획은 공휴일 텍스트가 있다면 이어서 쓰고, 없다면 강제 기록
                                existing = str(target_cell.value) if target_cell.value else ""
                                holiday_keywords = ["신정", "설날", "삼일절", "어린이날", "부처님오신날", "현충일", "광복절", "추석", "개천절", "한글날", "성탄절", "대체공휴일", "선거"]
                                if existing and len(existing) < 20 and any(h in existing for h in holiday_keywords):
                                    target_cell.value = f"{existing}\n{formatted_val}"
                                else:
                                    target_cell.value = formatted_val
                                target_cell.alignment = smart_align
                                target_cell.font = Font(size=font_size, bold=True, color=f_color)
                            else:
                                # 3순위: AI 수련내용은 빈칸일 때만 기록 (데이터 보호)
                                if not target_cell.value or (isinstance(target_cell.value, str) and len(target_cell.value) < 5):
                                    target_cell.value = formatted_val
                                    target_cell.alignment = smart_align
                                    target_cell.font = Font(size=font_size)
                        except: pass
                    else:
                        if is_annual:
                            target_machine_cell.value = formatted_val
                            target_machine_cell.alignment = smart_align
                            target_machine_cell.font = Font(size=font_size, bold=True, color=f_color)
                        else:
                            if not target_machine_cell.value:
                                target_machine_cell.value = formatted_val
                                target_machine_cell.alignment = smart_align
                                target_machine_cell.font = Font(size=font_size)
                    break

        wb.save(excel_path)
        print(f"✅ 월간 일정 {len(schedule_entries)}개가 엑셀에 등록되었습니다.")
        return True
    except Exception as e:
        print(f"❌ 엑셀 기록 오류: {e}")
        return False


def _init_annual_sheet(ws):
    """Initialize the 연간계획 sheet with proper headers."""
    ws.title = "연간계획"

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "🦁 라이온짐 연간 수련계획표"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="333333")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers
    headers = ["날짜", "행사명/수련주제", "장소", "수련단계(1-4)", "비고"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [15, 35, 20, 10, 15]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width


def _init_monthly_sheet(ws, year: int, month: int):
    """월간 달력 시트 초기화 (달력 그리드 + 하단 기계 판독 영역)"""
    ws.title = f"{month}월"
    
    # 상단 타이틀
    ws.merge_cells("A1:G1")
    ws["A1"] = f"🦁 {year}년 {month}월 라이온짐 수련계획표"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    # 정밀 공휴일 정보 가져오기
    holidays = get_holidays_for_month(year, month)

    # 요일 헤더
    days = ["일", "월", "화", "수", "목", "금", "토"]
    for i, day in enumerate(days, 1):
        cell = ws.cell(row=2, column=i, value=day)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if i == 1: cell.font = Font(bold=True, color="FF0000") # 일요일 빨강
        if i == 7: cell.font = Font(bold=True, color="0000FF") # 토요일 파랑

    # 달력 그리드 생성 (3~29행)
    calendar.setfirstweekday(calendar.SUNDAY)
    cal = calendar.monthcalendar(year, month)
    grid_map = {}
    
    current_row = 3
    for week in cal:
        # 날짜 행
        for i, day in enumerate(week, 1):
            if day != 0:
                cell = ws.cell(row=current_row, column=i, value=day)
                cell.alignment = Alignment(horizontal="left", vertical="top")
                
                # 일요일(i=1) 또는 공휴일이면 빨간색
                date_key = f"{year}-{month:02d}-{day:02d}"
                is_holiday = date_key in holidays
                if i == 1 or is_holiday:
                    cell.font = Font(bold=True, color="FF0000")
                else:
                    cell.font = Font(bold=True)
                
                grid_map[day] = cell.coordinate
        
        content_row = current_row + 1
        ws.row_dimensions[content_row].height = 60
        
        for i, day in enumerate(week, 1):
            if day != 0:
                content_cell = ws.cell(row=content_row, column=i)
                grid_map[day] = content_cell.coordinate
                
                # 공휴일이면 내용 칸에 공휴일 이름 적고 빨간색 처리
                date_key = f"{year}-{month:02d}-{day:02d}"
                if date_key in holidays:
                    content_cell.value = holidays[date_key]
                    content_cell.font = Font(bold=True, color="FF0000")
                    content_cell.alignment = Alignment(horizontal="center", vertical="center")
                
        current_row += 2 # 한 주당 2행 사용

    # --- 하단 기계 판독 영역 (31행~61행) ---
    ws.cell(row=31, column=1, value="날짜")
    ws.cell(row=31, column=2, value="1단계 (준비)")
    ws.cell(row=31, column=3, value="2단계 (체력)")
    ws.cell(row=31, column=4, value="3단계 (기술)")
    ws.cell(row=31, column=5, value="4단계 (마무리)")
    
    for day in range(1, 32):
        try:
            d_obj = date(year, month, day)
            r = 31 + day
            ws.cell(row=r, column=1, value=d_obj.strftime("%Y-%m-%d"))
            
            # 수식 연결 (그리드의 해당 날짜 셀을 바라보게 함)
            if day in grid_map:
                addr = grid_map[day]
                # 4단계 모두 일단 동일한 셀을 바라보게 초기화
                ws.cell(row=r, column=2, value=f"={addr}")
                ws.cell(row=r, column=3, value=f"={addr}")
                ws.cell(row=r, column=4, value=f"={addr}")
                ws.cell(row=r, column=5, value=f"={addr}")
        except ValueError:
            break # 해당 월의 마지막 날짜 지남

    # 열 너비 설정
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 15

    # --- 인쇄 설정 (가로 방향, 1페이지 꽉 차게, 달력 영역만 인쇄) ---
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    # 달력 그리드 영역까지만 인쇄 영역 지정 (A1 ~ G15)
    ws.print_area = "A1:G15"


def sync_annual_sheet_to_monthly_tabs(excel_path: str, weekend_events: list = None, skip_annual: bool = False) -> bool:
    """
    Read existing data from the '연간계획' sheet and spread it to all 12 monthly tabs.
    Also injects recurring weekend events from gym profile into empty saturdays.
    
    :param skip_annual: If True, only inject weekend_events and skip annual sheet.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl이 필요합니다. pip install openpyxl")

    try:
        wb = openpyxl.load_workbook(excel_path)
        entries = []
        if not skip_annual:
            if "연간계획" not in wb.sheetnames:
                print("⚠️ '연간계획' 시트가 없습니다. 연간계획 동기화를 건너뜁니다.")
            else:
                ws_annual = wb["연간계획"]
                for row in ws_annual.iter_rows(min_row=3, values_only=True):
                    if not row[0]: continue # No date
                    
                    d_val = row[0]
                    if isinstance(d_val, datetime):
                        d_str = d_val.strftime("%Y-%m-%d")
                    else:
                        d_str = str(d_val).strip()

                    entries.append({
                        "date": d_str,
                        "title": str(row[1]).strip() if row[1] else "",
                        "location": str(row[2]).strip() if row[2] else "",
                        "stage": str(row[3]).strip() if row[3] else "3"
                    })
                print(f"📊 연간계획 시트에서 {len(entries)}개의 항목을 읽었습니다.")
        else:
            print("ℹ️ 연간계획 동기화를 건너뛰고 주말 정기 행사만 처리합니다.")

        # 2. Group by month
        monthly_map = {}
        for entry in entries:
            date_str = entry.get("date", "")
            try:
                d_obj = None
                for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d"):
                    try:
                        d_obj = datetime.strptime(date_str, fmt)
                        break
                    except: continue
                
                if d_obj:
                    m = d_obj.month
                    entry["date"] = d_obj.strftime("%Y-%m-%d")
                    month_key = f"{m}월"
                    if month_key not in monthly_map: monthly_map[month_key] = []
                    monthly_map[month_key].append(entry)
            except: continue

        # 3. 1월~12월 순회하며 데이터 주입
        for m_idx in range(1, 13):
            month_name = f"{m_idx}월"
            if month_name not in wb.sheetnames: continue
            
            m_ws = wb[month_name]
            
            # 주말 행사 대기열 (밀린 행사 보관용)
            if m_idx == 1:
                pending_events = []

            # 인쇄 페이지 구분선 삽입
            existing_breaks = [b.id for b in m_ws.row_breaks.brk]
            if 30 not in existing_breaks:
                m_ws.row_breaks.append(Break(id=30))

            # [1순위] 연간계획 데이터 동기화 (대회, 협회 행사 등)
            if month_name in monthly_map:
                for entry in monthly_map[month_name]:
                    d_str = entry.get("date")
                    raw_title = entry.get("title")
                    if not raw_title: continue
                    
                    full_content = raw_title
                    if entry.get("location"):
                        full_content = f"{raw_title}\n({entry['location']})"
                    
                    stage_val = entry.get("stage", "3")
                    try:
                        target_col = int(re.search(r'\d', str(stage_val)).group())
                    except:
                        target_col = 3
                    
                    for row in m_ws.iter_rows(min_row=32, max_row=65):
                        cell_val = row[0].value
                        if cell_val and str(cell_val).startswith("="):
                            try: cell_val = m_ws[str(cell_val).replace("=","").strip()].value
                            except: pass
                        
                        c_date = cell_val.strftime("%Y-%m-%d") if isinstance(cell_val, datetime) else str(cell_val).strip()[:10].replace(".", "-")
                        
                        if c_date == d_str:
                            machine_cell = row[target_col]
                            if machine_cell.value and str(machine_cell.value).startswith("="):
                                t_addr = str(machine_cell.value).replace("=", "").strip()
                                try:
                                    t_cell = m_ws[t_addr]
                                    if not t_cell.value or len(str(t_cell.value).strip()) < 2:
                                        t_cell.value = full_content
                                        t_cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                                        t_cell.font = Font(size=10, bold=True, color="0000FF")
                                        print(f"  ✅ [연간동기화] {d_str} -> 달력본문({t_addr}) 기록 완료")
                                except: pass
                            break

            # [2순위] 체육관 정기 주말 행사 배치 (빈 토요일 대상)
            if weekend_events:
                year = 2026
                try:
                    match = re.search(r"20\d{2}", os.path.basename(excel_path))
                    if match: year = int(match.group())
                except: pass
                
                holidays = get_holidays_for_month(year, m_idx)
                
                # 가용 토요일 찾기
                available_saturdays = []
                total_saturdays = 0
                for row in m_ws.iter_rows(min_row=32, max_row=65):
                    cell_val = row[0].value
                    if not cell_val: continue
                    
                    # 날짜 값 추출 (수식인 경우 참조값 확인)
                    if cell_val and str(cell_val).startswith("="):
                        try: cell_val = m_ws[str(cell_val).replace("=","").strip()].value
                        except: pass
                    
                    # 날짜 객체로 변환 시도
                    if isinstance(cell_val, str):
                        try:
                            # '2026-01-01' 또는 '2026.01.01' 등 다양한 형식 시도
                            d_clean = cell_val.strip().replace(".", "-")[:10]
                            cell_val = datetime.strptime(d_clean, "%Y-%m-%d")
                        except: pass

                    if not isinstance(cell_val, datetime): continue
                    if cell_val.weekday() != 5: continue # 토요일(5)만 대상
                    
                    total_saturdays += 1
                    d_str = cell_val.strftime("%Y-%m-%d")
                    
                    # 공휴일 제외
                    if d_str in holidays:
                        print(f"    ☕ [건너뜀] {d_str} (공휴일)")
                        continue
                    
                    # 달력 본문의 해당 칸이 비어있는지 확인 (3단계/D열 기준)
                    # 감지용 셀(row[3])에 수식이 있다면 그 수식이 가리키는 실제 본문 셀 확인
                    machine_cell = row[3]
                    t_addr = None
                    if machine_cell.value and str(machine_cell.value).startswith("="):
                        t_addr = str(machine_cell.value).replace("=","").strip()
                    
                    if t_addr:
                        try:
                            t_cell = m_ws[t_addr]
                            # 본문 셀이 비어있거나 아주 짧은 경우에만 가용지로 판단
                            if not t_cell.value or len(str(t_cell.value).strip()) < 2:
                                available_saturdays.append({"date": d_str, "addr": t_addr})
                            else:
                                print(f"    ⚠️ [점유됨] {d_str} ({t_addr}: {str(t_cell.value)[:10]}...)")
                        except: pass
                    else:
                        # 수식이 없는 경우 감지용 셀 자체를 본문으로 간주 (비정상 케이스 대비)
                        if not machine_cell.value or len(str(machine_cell.value).strip()) < 2:
                            available_saturdays.append({"date": d_str, "addr": machine_cell.coordinate})

                # 1. 이번 달에 예정된 행사 대기열에 추가
                for ev in weekend_events:
                    cycle = int(ev.get("cycle", 1))
                    if (m_idx - 1) % cycle == 0:
                        name = ev.get("name", "").strip()
                        count = int(ev.get("count", 0))
                        if name and count > 0:
                            # 동일한 명칭의 행사가 이미 대기열에 있는지 확인 (중복 방지 원하면 조정 가능)
                            pending_events.append({"name": name, "count": count})

                if available_saturdays:
                    print(f"  🔍 [{month_name}] 토요일 {total_saturdays}개 중 {len(available_saturdays)}개 가용 확인 (대기열: {len(pending_events)}종)")
                    
                    # 2. 대기열에 있는 행사들을 가용 토요일에 배치
                    # (먼저 들어온 대기열 항목부터 순차적으로 배치)
                    temp_queue = []
                    while pending_events and available_saturdays:
                        curr_ev = pending_events.pop(0)
                        p_count = curr_ev["count"]
                        name = curr_ev["name"]
                        
                        placed_in_this_month = 0
                        for _ in range(p_count):
                            if not available_saturdays: break
                            target = available_saturdays.pop(0)
                            t_cell = m_ws[target["addr"]]
                            t_cell.value = f"[{name}]"
                            t_cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                            t_cell.font = Font(size=10, bold=True, color="FF6600")
                            placed_in_this_month += 1
                            print(f"    🟠 [정기배치] {target['date']} -> {target['addr']} ([{name}])")
                        
                        # 다 배치하지 못한 경우 남은 횟수를 기록하여 다시 대기열에 추가 (다음 달로 이월)
                        remaining = p_count - placed_in_this_month
                        if remaining > 0:
                            temp_queue.append({"name": name, "count": remaining})
                            print(f"    ⚠️ '{name}' {remaining}회 배정 실패 -> 다음 달로 이월됨")
                        else:
                            print(f"    ✅ '{name}' 배치 완료")
                    
                    # 배정 못한 나머지 대기열 유지
                    pending_events = temp_queue + pending_events
                else:
                    if pending_events:
                        print(f"  ❌ [{month_name}] 빈 토요일이 없어 {len(pending_events)}종의 행사가 모두 이월됩니다.")

        wb.save(excel_path)
        print(f"✅ 연간계획 및 정기 행사가 모든 월별 시트로 동기화되었습니다.")
        return True

    except Exception as e:
        print(f"❌ 동기화 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False


def reset_weekend_events(excel_path: str) -> bool:
    """
    엑셀 파일의 모든 월별 시트에서 정기 주말 행사(주황색 글씨)만 찾아서 삭제합니다.
    """
    if not OPENPYXL_AVAILABLE: return False
    if not os.path.exists(excel_path): return False

    try:
        wb = openpyxl.load_workbook(excel_path)
        count = 0
        for m_idx in range(1, 13):
            month_name = f"{m_idx}월"
            if month_name not in wb.sheetnames: continue
            
            m_ws = wb[month_name]
            # 감지용 셀 영역(32~65행)을 통해 달력 본문 셀 추적
            for row in m_ws.iter_rows(min_row=32, max_row=65):
                # 모든 단계(B,C,D,E열)에 대해 주말 행사가 있을 수 있으므로 체크
                for col_idx in range(1, 5): 
                    machine_cell = row[col_idx]
                    if machine_cell.value and str(machine_cell.value).startswith("="):
                        t_addr = str(machine_cell.value).replace("=","").strip()
                        try:
                            t_cell = m_ws[t_addr]
                            # 폰트 색상이 주황색(FF6600)인 경우만 정기 행사로 간주하여 삭제
                            if t_cell.font and t_cell.font.color and t_cell.font.color.rgb == "00FF6600":
                                t_cell.value = None
                                count += 1
                        except: pass
        
        wb.save(excel_path)
        print(f"🧹 주말 정기 행사 {count}개를 삭제했습니다.")
        return True
    except Exception as e:
        print(f"❌ 주말 행사 초기화 오류: {e}")
        return False


def reset_calendar_data(excel_path: str, month: int = None) -> bool:
    """
    엑셀 일정 파일에서 데이터를 삭제합니다.
    - month가 지정된 경우: 해당 월의 시트만 초기화
    - month가 None인 경우: 전체 연간계획 및 1~12월 초기화
    """
    if not OPENPYXL_AVAILABLE:
        print("⚠️ openpyxl이 필요합니다.")
        return False
    
    if not os.path.exists(excel_path):
        print(f"⚠️ 파일을 찾을 수 없습니다: {excel_path}")
        return False

    try:
        wb = openpyxl.load_workbook(excel_path)
        
        # 1. 연간계획 초기화 (전체 초기화 모드일 때만)
        if month is None and "연간계획" in wb.sheetnames:
            ws = wb["연간계획"]
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

        # 2. 월간 달력 초기화
        target_months = [f"{month}월"] if month else [f"{i}월" for i in range(1, 13)]
        
        for m_name in target_months:
            if m_name in wb.sheetnames:
                ws = wb[m_name]
                # 하단 기계 판독 영역(31행~)의 수식을 따라가서 실제 달력 칸 삭제
                for row in ws.iter_rows(min_row=31, max_row=ws.max_row):
                    for col_idx in [1, 2, 3, 4]: # 1~4단계 전체
                        cell = row[col_idx]
                        if cell.value and str(cell.value).startswith("="):
                            target_addr = str(cell.value).replace("=", "").strip()
                            try:
                                ws[target_addr].value = None
                            except: pass
                
                # 수동으로 입력된 칸들도 비움 (수식이나 날짜 제외 텍스트만)
                for row in ws.iter_rows(min_row=3, max_row=30):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and not cell.value.startswith("="):
                            cell.value = None
                
                # [중요] 초기화 후 요일 헤더 및 공휴일 명칭/색상 즉시 복구
                year = 2026
                try:
                    match = re.search(r"20\d{2}", os.path.basename(excel_path))
                    if match: year = int(match.group())
                except: pass
                fix_layout_and_holidays(ws, year, int(m_name.replace("월","")))

        wb.save(excel_path)
        print(f"✅ '{excel_path}' 초기화 완료 (헤더 및 공휴일 보존)")
        return True
    except Exception as e:
        print(f"❌ 초기화 중 오류: {e}")
        return False
