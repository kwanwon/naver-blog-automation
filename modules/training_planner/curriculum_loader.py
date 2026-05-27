# -*- coding: utf-8 -*-
"""
curriculum_loader.py
사용자가 업로드한 다양한 파일(PDF, TXT, 엑셀, 이미지)을 읽어
AI가 최적으로 소화할 수 있는 Markdown 형식(curriculum_data.md)으로 변환·저장합니다.

[완전 독립 모듈 - 기존 자동화 시스템에 영향 없음]
"""

import os
import json
import re
from datetime import datetime


# --- 선택적 임포트 (없어도 기본 기능은 동작) ---
try:
    import pdfplumber
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ── 기본 경로 ──────────────────────────────────────────────
BASE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "config", "training_planner")
CURRICULUM_MD_PATH = os.path.join(BASE_DIR, "curriculum_data.md")
GYM_PROFILE_PATH = os.path.join(BASE_DIR, "gym_profile.json")
AGE_CATEGORIES_PATH = os.path.join(BASE_DIR, "age_categories.json")


def ensure_dirs():
    """설정 폴더가 없으면 생성합니다."""
    os.makedirs(BASE_DIR, exist_ok=True)


# ─────────────────────────────────────────────────────────
# 체육관 프로필 관리
# ─────────────────────────────────────────────────────────

def load_gym_profile() -> dict:
    """저장된 체육관 프로필을 불러옵니다. 없으면 기본값을 반환합니다."""
    ensure_dirs()
    if os.path.exists(GYM_PROFILE_PATH):
        try:
            with open(GYM_PROFILE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    # 기본 프로필
    return {
        "gym_name": "",
        "sport": "",
        "concept": "",
        "concept_detail": "",
        "instructor_name": "",
        "routine_1": "몸풀기, 스트레칭, 기본기",
        "routine_2": "팔벌려뛰기, 근력운동",
        "routine_3": "낙법, 호신술, 대련전술",
        "routine_4": "정리운동, 레크레이션",
        "weekend_events": [
            {"name": "특강", "cycle": 1, "count": 1},
            {"name": "캠프", "cycle": 1, "count": 0},
            {"name": "합숙", "cycle": 3, "count": 0},
            {"name": "", "cycle": 1, "count": 0},
            {"name": "", "cycle": 1, "count": 0}
        ],
        "last_updated": ""
    }


def save_gym_profile(profile: dict) -> bool:
    """체육관 프로필을 저장합니다."""
    ensure_dirs()
    try:
        profile["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(GYM_PROFILE_PATH, "w", encoding="utf-8") as f:
            json.dump(profile, f, ensure_ascii=False, indent=2)
        print(f"✅ [체육관 프로필] 저장 완료: {GYM_PROFILE_PATH}")
        return True
    except Exception as e:
        print(f"❌ [체육관 프로필] 저장 실패: {e}")
        return False


# ─────────────────────────────────────────────────────────
# 연령별 카테고리 관리
# ─────────────────────────────────────────────────────────

DEFAULT_AGE_CATEGORIES = [
    {
        "id": "infant",
        "name": "🧒 유치부",
        "description": "5~7세 (유아)",
        "age_range": "5~7",
        "training_style": "놀이 중심, 기초 체조, 집중력 게임, 15분 단위 활동 전환",
        "ratio": {"play": 50, "basic": 30, "physical": 20},
        "session_duration": 40
    },
    {
        "id": "elem_low",
        "name": "📚 초등 저학년",
        "description": "1~3학년 (초등 저)",
        "age_range": "7~10",
        "training_style": "기본기 + 레크레이션, 성장 스트레칭, 순발력 게임",
        "ratio": {"basic": 30, "technique": 30, "play": 40},
        "session_duration": 50
    },
    {
        "id": "elem_high",
        "name": "💪 초등 고학년",
        "description": "4~6학년 (초등 고)",
        "age_range": "10~13",
        "training_style": "기술 심화, 약속 대련, 체력 훈련, 대회 준비",
        "ratio": {"technique": 40, "physical": 30, "sparring": 30},
        "session_duration": 60
    },
    {
        "id": "middle_up",
        "name": "🔥 중등부 이상",
        "description": "중학생 이상 (청소년·선수부)",
        "age_range": "13+",
        "training_style": "실전 대련, 고급 기술, 심판 교육, 리더십 훈련",
        "ratio": {"sparring": 50, "physical": 30, "theory": 20},
        "session_duration": 70
    },
    {
        "id": "adult",
        "name": "🏋️ 성인·노인",
        "description": "성인 및 시니어",
        "age_range": "18+",
        "training_style": "건강 관리, 관절 보호, 낙상 예방, 단전호흡 위주",
        "ratio": {"health": 50, "basic": 30, "breathing": 20},
        "session_duration": 60
    }
]


def load_age_categories() -> list:
    """연령별 카테고리를 불러옵니다."""
    ensure_dirs()
    if os.path.exists(AGE_CATEGORIES_PATH):
        try:
            with open(AGE_CATEGORIES_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return DEFAULT_AGE_CATEGORIES


def save_age_categories(categories: list) -> bool:
    """연령별 카테고리를 저장합니다."""
    ensure_dirs()
    try:
        with open(AGE_CATEGORIES_PATH, "w", encoding="utf-8") as f:
            json.dump(categories, f, ensure_ascii=False, indent=2)
        print(f"✅ [연령 카테고리] 저장 완료")
        return True
    except Exception as e:
        print(f"❌ [연령 카테고리] 저장 실패: {e}")
        return False


# ─────────────────────────────────────────────────────────
# 파일 → Markdown 변환 (핵심 학습 데이터 생성)
# ─────────────────────────────────────────────────────────

def extract_text_from_file(file_path: str) -> str:
    """
    파일 형식을 자동 감지하여 텍스트를 추출합니다.
    지원 형식: TXT, PDF, XLSX/XLS
    """
    ext = os.path.splitext(file_path)[1].lower()
    print(f"[Step 1] [커리큘럼 로더] 파일 읽기 시작: {os.path.basename(file_path)} (상태: 시도)")

    if ext == ".txt":
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
            print(f"[Step 1] [커리큘럼 로더] TXT 읽기 (상태: 성공) - {len(text)}자")
            return text
        except Exception as e:
            print(f"[Step 1] [커리큘럼 로더] TXT 읽기 (상태: 실패) - {e}")
            return ""

    elif ext == ".pdf":
        if not PDF_AVAILABLE:
            print("⚠️ pdfplumber가 설치되지 않았습니다. (pip install pdfplumber)")
            return ""
        try:
            texts = []
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        texts.append(t)
            result = "\n".join(texts)
            print(f"[Step 1] [커리큘럼 로더] PDF 읽기 (상태: 성공) - {len(result)}자")
            return result
        except Exception as e:
            print(f"[Step 1] [커리큘럼 로더] PDF 읽기 (상태: 실패) - {e}")
            return ""

    elif ext in [".xlsx", ".xls"]:
        if not EXCEL_AVAILABLE:
            print("⚠️ openpyxl이 설치되지 않았습니다. (pip install openpyxl)")
            return ""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"## 시트: {sheet}")
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if row_vals:
                        lines.append(" | ".join(row_vals))
            result = "\n".join(lines)
            print(f"[Step 1] [커리큘럼 로더] 엑셀 읽기 (상태: 성공) - {len(result)}자")
            return result
        except Exception as e:
            print(f"[Step 1] [커리큘럼 로더] 엑셀 읽기 (상태: 실패) - {e}")
            return ""

    else:
        print(f"⚠️ 지원하지 않는 파일 형식입니다: {ext}")
        return ""


def learn_from_files(file_paths: list, gym_profile: dict) -> dict:
    """
    여러 파일에서 텍스트를 추출하고 curriculum_data.md에 저장합니다.
    반환값: {"success": bool, "item_count": int, "summary": str}
    """
    ensure_dirs()
    all_texts = []

    for path in file_paths:
        if not os.path.exists(path):
            print(f"⚠️ 파일이 존재하지 않습니다: {path}")
            continue
        text = extract_text_from_file(path)
        if text.strip():
            fname = os.path.basename(path)
            all_texts.append(f"### 파일: {fname}\n\n{text}\n")

    if not all_texts:
        return {"success": False, "item_count": 0, "summary": "추출된 텍스트가 없습니다."}

    combined = "\n---\n".join(all_texts)

    # Markdown 파일로 저장
    gym_name = gym_profile.get("gym_name", "체육관")
    sport = gym_profile.get("sport", "종목")
    header = f"""# {gym_name} {sport} 수련 커리큘럼 데이터
> 생성일: {datetime.now().strftime("%Y-%m-%d %H:%M")}
> AI 학습용 자료 - 사용자 업로드 기반

"""
    full_md = header + combined

    try:
        with open(CURRICULUM_MD_PATH, "w", encoding="utf-8") as f:
            f.write(full_md)

        # 항목 수 대략 계산 (줄 수 기반)
        item_count = len([l for l in full_md.split("\n") if l.strip()])
        kb_size = round(len(full_md.encode("utf-8")) / 1024, 1)

        print(f"✅ [커리큘럼 로더] 학습 데이터 저장 완료: {CURRICULUM_MD_PATH} ({kb_size}KB, 약 {item_count}줄)")
        return {
            "success": True,
            "item_count": item_count,
            "kb_size": kb_size,
            "summary": f"{len(file_paths)}개 파일에서 약 {item_count}개 항목 학습 완료 ({kb_size}KB)"
        }
    except Exception as e:
        print(f"❌ [커리큘럼 로더] 저장 실패: {e}")
        return {"success": False, "item_count": 0, "summary": str(e)}


def load_curriculum_md() -> str:
    """저장된 커리큘럼 Markdown 데이터를 불러옵니다."""
    if os.path.exists(CURRICULUM_MD_PATH):
        try:
            with open(CURRICULUM_MD_PATH, "r", encoding="utf-8") as f:
                return f.read()
        except Exception as e:
            print(f"⚠️ 커리큘럼 데이터 로드 실패: {e}")
    return ""


def get_curriculum_status() -> dict:
    """현재 학습 데이터의 상태 정보를 반환합니다."""
    if not os.path.exists(CURRICULUM_MD_PATH):
        return {"exists": False, "item_count": 0, "kb_size": 0, "last_updated": "없음"}
    try:
        with open(CURRICULUM_MD_PATH, "r", encoding="utf-8") as f:
            content = f.read()
        item_count = len([l for l in content.split("\n") if l.strip()])
        kb_size = round(len(content.encode("utf-8")) / 1024, 1)
        mtime = os.path.getmtime(CURRICULUM_MD_PATH)
        last_updated = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
        return {
            "exists": True,
            "item_count": item_count,
            "kb_size": kb_size,
            "last_updated": last_updated
        }
    except Exception:
        return {"exists": False, "item_count": 0, "kb_size": 0, "last_updated": "오류"}
