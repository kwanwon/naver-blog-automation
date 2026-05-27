# -*- coding: utf-8 -*-
"""
calendar_writer.py
AI가 생성한 수련계획 JSON 데이터를 엑셀 파일의 월간 달력에 기록합니다.
4단계 수련 루틴(1~4단계)을 엑셀의 기계 판독 영역에 각각 분리하여 기록합니다.

[수련계획 AI 시스템 전용 - 최신 버전]
"""

import os
from datetime import datetime
from modules.schedule_parser import fix_layout_and_holidays

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def write_plan_to_excel(excel_path: str, plan_entries: list) -> dict:
    """
    AI가 생성한 수련계획을 엑셀 월간 달력의 빈칸에만 기록합니다.
    4단계 루틴을 각각의 컬럼(1~4단계)에 분리하여 저장합니다.
    """
    if not OPENPYXL_AVAILABLE:
        return {"success": False, "written": 0, "skipped": 0, "error": "openpyxl 없음"}

    if not os.path.exists(excel_path):
        return {"success": False, "written": 0, "skipped": 0, "error": f"파일 없음: {excel_path}"}

    if not plan_entries:
        return {"success": False, "written": 0, "skipped": 0, "error": "기록할 데이터가 없습니다"}

    print(f"[Step 2] [수련계획 AI] 엑셀 기록 시작: {os.path.basename(excel_path)}")

    try:
        wb = openpyxl.load_workbook(excel_path)
        written = 0
        skipped = 0

        # 날짜별로 그룹화
        by_month = {}
        for entry in plan_entries:
            date_str = entry.get("date", "")
            title = entry.get("title", "").strip()
            if not date_str or not title:
                continue
            try:
                d = datetime.strptime(date_str, "%Y-%m-%d")
                month_key = f"{d.month}월"
                if month_key not in by_month:
                    by_month[month_key] = []
                by_month[month_key].append(entry)
            except ValueError:
                continue

        for month_name, entries in by_month.items():
            if month_name not in wb.sheetnames:
                print(f"⚠️ [수련계획 AI] '{month_name}' 탭이 없어 건너뜁니다.")
                continue

            ws = wb[month_name]

            for entry in entries:
                date_str = entry["date"]
                title = entry["title"].strip()

                # 날짜 행 찾기 (31행 이후 기계 판독 영역)
                matched = False
                for row in ws.iter_rows(min_row=31, max_row=65):
                    cell_val = row[0].value
                    
                    # 수식 연결 해결
                    if cell_val and str(cell_val).startswith("="):
                        try:
                            addr = str(cell_val).replace("=", "").strip()
                            cell_val = ws[addr].value
                        except: pass

                    if not cell_val: continue
                        
                    if isinstance(cell_val, datetime):
                        cell_date_str = cell_val.strftime("%Y-%m-%d")
                    else:
                        cell_date_str = str(cell_val).strip()[:10].replace(".", "-")
                    
                    # 날짜 매칭 성공 시
                    if cell_date_str == date_str or cell_date_str.endswith(date_str[5:]):
                        # AI 생성 내용을 4단계 루틴으로 분리 (\n 기준)
                        lines = [line.strip() for line in title.split('\n') if line.strip()]
                        
                        # 각 단계를 1단계(B열) ~ 4단계(E열)에 분산 기록
                        for i, line in enumerate(lines[:4]):
                            col_idx = i + 1 # B=1, C=2, D=3, E=4 (row 인덱스 기준)
                            target_machine_cell = row[col_idx]
                            
                            # 수식 참조 셀(달력 칸) 찾기
                            if target_machine_cell.value and str(target_machine_cell.value).startswith("="):
                                addr = str(target_machine_cell.value).replace("=", "").strip()
                                try:
                                    real_cell = ws[addr]
                                    # 빈칸일 때만 기록 (데이터 보호)
                                    if not real_cell.value or not str(real_cell.value).strip():
                                        if i == 0: # 첫 번째 단계에서 전체 내용을 달력 칸에 기록
                                            real_cell.value = title
                                            real_cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                                            real_cell.font = Font(size=_calc_font_size(title), bold=True)
                                        
                                        # 기계 판독 영역 해당 단계 열에도 기록
                                        target_machine_cell.value = line
                                        written += 1
                                        matched = True
                                    else:
                                        skipped += 1
                                except: pass
                            else:
                                # 수식이 없더라도 빈칸이면 기록
                                if not target_machine_cell.value or not str(target_machine_cell.value).strip():
                                    target_machine_cell.value = line
                                    written += 1
                                    matched = True
                                else:
                                    skipped += 1
                        break

            # 디자인 적용 (디테일 스타일, 테두리, 이모지 등)
            try:
                # 시트 이름(예: '5월')에서 숫자만 추출
                m_num = int(month_name.replace("월", ""))
                # 현재 연도 추출 (계획 데이터의 날짜 기준)
                y_num = datetime.strptime(entries[0]["date"], "%Y-%m-%d").year
                fix_layout_and_holidays(ws, y_num, m_num)
            except: pass

        wb.save(excel_path)
        print(f"✅ [수련계획 AI] 기록 완료 - 신규: {written}개, 스킵: {skipped}개")
        return {"success": True, "written": written, "skipped": skipped}

    except Exception as e:
        print(f"❌ [수련계획 AI] 오류: {e}")
        return {"success": False, "written": 0, "skipped": 0, "error": str(e)}


def get_existing_events_from_excel(excel_path: str, month_num: int) -> list:
    """엑셀에서 기존 일정 정보를 읽어옵니다."""
    if not OPENPYXL_AVAILABLE or not os.path.exists(excel_path):
        return []

    try:
        wb = openpyxl.load_workbook(excel_path)
        month_name = f"{month_num}월"
        if month_name not in wb.sheetnames: return []

        ws = wb[month_name]
        existing = []

        for row in ws.iter_rows(min_row=31, max_row=65):
            date_cell = row[0].value
            if date_cell and str(date_cell).startswith("="):
                try: date_cell = ws[str(date_cell).replace("=","").strip()].value
                except: pass

            if not date_cell: continue
            d_str = date_cell.strftime("%Y-%m-%d") if isinstance(date_cell, datetime) else str(date_cell).strip()[:10].replace(".", "-")

            # 1~4단계 중 하나라도 내용이 있으면 기존 일정으로 간주
            for col_idx in [1, 2, 3, 4]:
                val = row[col_idx].value
                if val and str(val).startswith("="):
                    try: val = ws[str(val).replace("=","").strip()].value
                    except: pass
                
                if val and str(val).strip() and not str(val).startswith("="):
                    existing.append({"date": d_str, "title": str(val).strip()})
                    break
        return existing
    except: return []


def _calc_font_size(text: str) -> int:
    """
    관장님 요청 사항 반영:
    - 기본 10포인트 (20자까지 한 줄 표기 가능)
    - 범위: 9 ~ 11포인트
    - 글자 수에 따른 자동 조절
    """
    if not text: return 10
    length = len(text)
    
    if length <= 12:
        return 11  # 아주 짧은 경우 가독성을 위해 11
    elif length <= 22:
        return 10  # 20자 내외(공백 포함 22자까지 여유)는 기준인 10
    else:
        return 9   # 그 이상 길어지는 경우 9로 축소하여 칸 넘침 방지
